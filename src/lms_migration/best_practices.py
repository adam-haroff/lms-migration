from __future__ import annotations

import argparse
import csv
import difflib
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


@dataclass(frozen=True)
class PracticeRow:
    row_number: int
    row_id: str
    source: str
    recommendation: str
    target: str
    notes: str
    category: str


HEADER_ALIASES = {
    "id": {"id", "ruleid", "recommendationid", "guidelineid", "key"},
    "source": {"source", "d2lfeature", "from", "origin", "currentstate", "item"},
    "recommendation": {
        "recommendation",
        "bestpractice",
        "action",
        "suggestedaction",
        "guidance",
        "recommendedaction",
    },
    "target": {"target", "canvasequivalent", "to", "canvasmapping", "destination"},
    "notes": {"notes", "details", "rationale", "accessibility", "wcag", "comments"},
    "category": {"category", "type", "domain", "module", "contenttype"},
}


def _normalize_header(header: str) -> str:
    return re.sub(r"[^a-z0-9]", "", header.strip().lower())


def _normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower())


def _find_header(headers: Iterable[str], aliases: set[str]) -> str | None:
    normalized_map = {header: _normalize_header(header) for header in headers if header}
    for header, normalized in normalized_map.items():
        if normalized in aliases:
            return header
    return None


def _pick_columns(headers: list[str]) -> dict[str, str | None]:
    return {
        key: _find_header(headers, aliases)
        for key, aliases in HEADER_ALIASES.items()
    }


def _get_cell(row: dict[str, str], key: str | None) -> str:
    if key is None:
        return ""
    return str(row.get(key, "") or "").strip()


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        return [dict(row) for row in reader]


def _read_xlsx(path: Path, sheet_name: str | None) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook[workbook.sheetnames[0]]

    rows = worksheet.iter_rows(values_only=True)
    try:
        headers_row = next(rows)
    except StopIteration:
        return []

    headers = [str(value).strip() if value is not None else "" for value in headers_row]
    output: list[dict[str, str]] = []

    for row_values in rows:
        row_dict: dict[str, str] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = row_values[index] if index < len(row_values) else ""
            row_dict[header] = str(value).strip() if value is not None else ""
        output.append(row_dict)

    return output


def _read_table(path: Path, sheet_name: str | None) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix == ".xlsx":
        return _read_xlsx(path, sheet_name)
    raise ValueError(f"Unsupported file type: {suffix}. Use .csv or .xlsx")


def _build_rows(raw_rows: list[dict[str, str]]) -> list[PracticeRow]:
    if not raw_rows:
        return []

    headers = list(raw_rows[0].keys())
    columns = _pick_columns(headers)
    rows: list[PracticeRow] = []

    for index, raw in enumerate(raw_rows, start=2):
        source = _get_cell(raw, columns["source"])
        recommendation = _get_cell(raw, columns["recommendation"])
        target = _get_cell(raw, columns["target"])
        notes = _get_cell(raw, columns["notes"])
        row_id = _get_cell(raw, columns["id"])
        category = _get_cell(raw, columns["category"])

        if not any((source, recommendation, target, notes, row_id, category)):
            continue

        rows.append(
            PracticeRow(
                row_number=index,
                row_id=row_id,
                source=source,
                recommendation=recommendation,
                target=target,
                notes=notes,
                category=category,
            )
        )

    return rows


def _fingerprint(row: PracticeRow) -> str:
    parts = [row.source, row.recommendation, row.target, row.notes, row.category]
    return " | ".join(_normalize_text(part) for part in parts)


def _find_duplicates(rows: list[PracticeRow]) -> list[list[PracticeRow]]:
    grouped: dict[str, list[PracticeRow]] = defaultdict(list)
    for row in rows:
        grouped[_fingerprint(row)].append(row)
    return [group for group in grouped.values() if len(group) > 1]


def _is_negative_statement(text: str) -> bool:
    lowered = _normalize_text(text)
    negative_markers = ["do not", "don't", "avoid", "never", "remove", "disable"]
    return any(marker in lowered for marker in negative_markers)


def _find_conflicts(rows: list[PracticeRow]) -> list[dict[str, object]]:
    grouped: dict[str, list[PracticeRow]] = defaultdict(list)
    for row in rows:
        key = _normalize_text(row.source or row.row_id)
        if key:
            grouped[key].append(row)

    conflicts: list[dict[str, object]] = []
    for key, group in grouped.items():
        recommendations = {
            _normalize_text(f"{row.recommendation} {row.target}"): row for row in group if row.recommendation or row.target
        }
        if len(recommendations) <= 1:
            continue

        polarity = {_is_negative_statement(f"{row.recommendation} {row.target}") for row in group}
        if len(polarity) > 1 or len(recommendations) > 1:
            conflicts.append(
                {
                    "source_key": key,
                    "rows": [
                        {
                            "row_number": row.row_number,
                            "row_id": row.row_id,
                            "source": row.source,
                            "recommendation": row.recommendation,
                            "target": row.target,
                        }
                        for row in group
                    ],
                }
            )

    return conflicts


def _candidate_similarity_rows(rows: list[PracticeRow]) -> list[tuple[PracticeRow, PracticeRow, float]]:
    pairs: list[tuple[PracticeRow, PracticeRow, float]] = []

    if len(rows) > 1200:
        return pairs

    for i, left in enumerate(rows):
        left_text = _normalize_text(f"{left.source} {left.recommendation} {left.target}")
        if not left_text:
            continue
        for right in rows[i + 1 :]:
            right_text = _normalize_text(f"{right.source} {right.recommendation} {right.target}")
            if not right_text:
                continue
            score = difflib.SequenceMatcher(a=left_text, b=right_text).ratio()
            if score >= 0.9 and _fingerprint(left) != _fingerprint(right):
                pairs.append((left, right, score))

    return pairs


def _report_markdown(report: dict) -> str:
    lines = [
        "# Best Practices Audit Report",
        "",
        f"Input: `{report['input']}`",
        f"Rows evaluated: {report['summary']['rows_evaluated']}",
        "",
        "## Findings",
        "",
        f"- Exact duplicates: {report['summary']['exact_duplicates']}",
        f"- Potential conflicts: {report['summary']['potential_conflicts']}",
        f"- Potential redundancies: {report['summary']['potential_redundancies']}",
        "",
    ]

    if report["duplicates"]:
        lines.append("## Duplicate Groups")
        lines.append("")
        for group in report["duplicates"]:
            row_numbers = ", ".join(str(item["row_number"]) for item in group)
            lines.append(f"- Rows: {row_numbers}")
        lines.append("")

    if report["conflicts"]:
        lines.append("## Conflict Groups")
        lines.append("")
        for conflict in report["conflicts"]:
            row_numbers = ", ".join(str(item["row_number"]) for item in conflict["rows"])
            lines.append(f"- Source key `{conflict['source_key']}` in rows: {row_numbers}")
        lines.append("")

    if report["redundancies"]:
        lines.append("## Redundancy Candidates")
        lines.append("")
        for item in report["redundancies"][:25]:
            lines.append(
                f"- Rows {item['left_row']} and {item['right_row']} (similarity: {item['similarity']})"
            )
        lines.append("")

    return "\n".join(lines)


def run_audit(input_path: Path, output_dir: Path, sheet_name: str | None = None) -> tuple[Path, Path]:
    raw_rows = _read_table(input_path, sheet_name)
    rows = _build_rows(raw_rows)

    duplicates = _find_duplicates(rows)
    conflicts = _find_conflicts(rows)
    redundancies = _candidate_similarity_rows(rows)

    report = {
        "input": str(input_path),
        "summary": {
            "rows_evaluated": len(rows),
            "exact_duplicates": sum(len(group) - 1 for group in duplicates),
            "potential_conflicts": len(conflicts),
            "potential_redundancies": len(redundancies),
        },
        "duplicates": [
            [
                {
                    "row_number": row.row_number,
                    "row_id": row.row_id,
                    "source": row.source,
                    "recommendation": row.recommendation,
                    "target": row.target,
                }
                for row in group
            ]
            for group in duplicates
        ],
        "conflicts": conflicts,
        "redundancies": [
            {
                "left_row": left.row_number,
                "right_row": right.row_number,
                "similarity": round(score, 3),
                "left_text": f"{left.source} | {left.recommendation} | {left.target}",
                "right_text": f"{right.source} | {right.recommendation} | {right.target}",
            }
            for left, right, score in redundancies
        ],
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    stem = input_path.stem
    json_path = output_dir / f"{stem}.best-practices-audit.json"
    md_path = output_dir / f"{stem}.best-practices-audit.md"

    json_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    md_path.write_text(_report_markdown(report), encoding="utf-8")

    return json_path, md_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="lms-best-practices-audit",
        description="Audit migration best-practices spreadsheet for duplicate/conflicting guidance",
    )
    parser.add_argument("input", type=Path, help="Path to .xlsx or .csv best-practices file")
    parser.add_argument("--sheet", type=str, default=None, help="Optional worksheet name (.xlsx only)")
    parser.add_argument("--output-dir", type=Path, default=Path("output"), help="Output report directory")
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    if not args.input.exists():
        parser.error(f"Input file not found: {args.input}")

    try:
        json_path, md_path = run_audit(
            input_path=args.input,
            output_dir=args.output_dir,
            sheet_name=args.sheet,
        )
    except Exception as exc:
        parser.error(str(exc))

    print(f"Audit JSON report: {json_path}")
    print(f"Audit Markdown report: {md_path}")


if __name__ == "__main__":
    main()
